import os
import pathlib
import tempfile
import warnings
import re
import itertools
from collections import OrderedDict
import time

import numpy as np

import matplotlib
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import plotnine as p9

import rdkit
import rdkit.Chem
import rdkit.Chem.Draw
from rdkit.Chem.Draw import IPythonConsole

# from rdkit import Chem
# from rdkit.Chem import Draw
# from rdkit.Chem.MolStandardize import rdMolStandardize

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from tqdm import tqdm

import natsort


def scale_dimensions_to_fit(original_width, original_height, max_width, max_height):
    """
    Scales (width, height) to fit within max dimensions, preserving aspect ratio.

    Returns: (new_width, new_height)
    """

    asp_ratio = original_width / original_height

    new_width = max_width
    new_height = max_width / asp_ratio

    if new_height > max_height:
        new_height = max_height
        new_width = max_height * asp_ratio

    return int(new_width), int(new_height)


def write_to_excel_cell(filename, sheet_name, texts, images=None, column_width=None, row_height=None, img_scale_fact=1.33):
    # Create a new workbook if file doesn't exist, else load it workbook if file doesn't exist, else load it
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    # Get or create the desired sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    for cell, content in texts.items():
        ws[cell] = content
        col_letter = "".join(filter(str.isalpha, cell))
        row_number = int("".join(filter(str.isdigit, cell)))
        if column_width is not None:
            ws.column_dimensions[col_letter].width = column_width
        if row_height is not None:
            ws.row_dimensions[row_number].height = row_height

    if images is not None:
        for cell, image in images.items():
            if not os.path.exists(image):
                raise FileNotFoundError(f"Image file '{image}' not found.")
            # Load image and add to sheetimg_scale_factimg_scale_fact
            img = ExcelImage(image)
            img.width, img.height = scale_dimensions_to_fit(img.width, img.height, column_width * img_scale_fact, row_height * img_scale_fact)

            # Position image to the top-left of the cell
            ws.add_image(img, cell)

            # Optional: Adjust column width and row height
            col_letter = "".join(filter(str.isalpha, cell))
            row_number = int("".join(filter(str.isdigit, cell)))
            ws.column_dimensions[col_letter].width = column_width
            ws.row_dimensions[row_number].height = row_height

    wb.save(filename)


def write_table_to_excel(filename, table, sheet_name="Sheet", img_prefix="$$$IMG:", column_width=40, row_height=8):
    """
    Writes a table (list of lists) to an Excel file using the write_to_excel_cell function.
    Deletes the file first if it exists.

    Args:
        filename (str): The path to the Excel file.
        table (list of list): The data to write, where each inner list is a row.
        sheet_name (str, optional): The name of the sheet to write to. Defaults to "Sheet1".

    Raises:
        OSError: If the existing file cannot be deleted.
    """
    if os.path.exists(filename):
        try:
            os.remove(filename)
        except OSError as e:
            raise OSError(f"Error: Could not delete existing file '{filename}'. Reason: {e}")

    texts_to_write = {}
    images_to_write = {}
    for r_idx, row_data in enumerate(table):
        for c_idx, cell_value in enumerate(row_data):
            # Convert 0-based c_idx to 1-based column letter
            col_letter = get_column_letter(c_idx + 1)
            # Excel rows are 1-based
            cell_address = f"{col_letter}{r_idx + 1}"

            if isinstance(cell_value, str) and cell_value.startswith(img_prefix):
                image_path = cell_value.replace(img_prefix, "", 1)
                # Ensure 'images_to_write' is initialized in the function scope, e.g., images_to_write = {}
                images_to_write[cell_address] = image_path
                # Set cell_value to an empty string so the subsequent line doesn't write the image directive as text
            else:
                # If it's not an image string, cell_value remains unchanged,
                # and the original `texts_to_write[cell_address] = cell_value` line will handle it as text.
                texts_to_write[cell_address] = cell_value

    # It will create a new workbook since we've deleted the file if it existed.
    write_to_excel_cell(filename, sheet_name, texts_to_write, images=images_to_write, column_width=column_width, row_height=row_height)


def list_to_excel_table(data, filename, sheet_name="Sheet", img_prefix="$$$IMG:", column_width=40, row_height=8):
    """
    Converts a list of dictionaries into a table and writes it to an Excel file.

    The input list contains the rows and their data. The keys of these
    inner dictionaries are used as column headers.

    Args:
        data_dict (list): A list of dictionaries mapping column headers (strings)
                            to cell values.
                            Example: {"Row1": {"ColA": 1, "ColB": "img_path.png"}, "Row2": {"ColA": 2}}
        filename (str): The path to the Excel file to be created/overwritten.
        sheet_name (str, optional): The name of the sheet to write to. Defaults to "Sheet".
        row_key_col_name (str, optional): The header for the first column, which will
                                            contain the row keys from data_dict. Defaults to "Row_Key".
        img_prefix (str, optional): The prefix used to identify image paths in cell values.
                                    Defaults to "$$$IMG:".
    """
    if not isinstance(data, list):
        raise TypeError("Input datalist must be a list.")

    # Extract all unique keys from the inner dictionaries to form column headers
    unique_column_headers = set()
    if data:
        for inner_dict in data:
            if isinstance(inner_dict, dict):
                unique_column_headers.update(inner_dict.keys())
            # If inner_dict is not a dict, its keys won't be added.
            # This means columns are only derived from valid inner dictionaries.
        all_column_headers = natsort.natsorted(list(unique_column_headers))

        # Prepare the table (list of lists)
        table_data = []

        # Add the header row
        header_row = all_column_headers
        table_data.append(header_row)

        # Add the data rows
        for inner_dict in data:
            current_row_values = []
            if isinstance(inner_dict, dict):
                for col_header in all_column_headers:
                    current_row_values.append(inner_dict.get(col_header, ""))  # Use empty string for missing values
            else:
                # If the value for a row_key is not a dict, fill data cells with empty strings
                current_row_values.extend([""] * len(all_column_headers))
            table_data.append(current_row_values)

        # Write the table to Excel using the provided function
        write_table_to_excel(filename, table_data, sheet_name=sheet_name, img_prefix=img_prefix, column_width=column_width, row_height=row_height)


def standardize_smile(smiles):
    # follows the steps in
    # https://github.com/greglandrum/RSC_OpenScience_Standardization_202104/blob/main/MolStandardize%20pieces.ipynb
    # as described **excellently** (by Greg) in
    # https://www.youtube.com/watch?v=eWTApNX8dJQ
    mol = rdkit.Chem.MolFromSmiles(smiles)

    # removeHs, disconnect metal atoms, normalize the molecule, reionize the molecule
    clean_mol = rdMolStandardize.Cleanup(mol)

    # if many fragments, get the "parent" (the actual mol we are interested in)
    parent_clean_mol = rdMolStandardize.FragmentParent(clean_mol)

    # try to neutralize molecule
    uncharger = rdMolStandardize.Uncharger()  # annoying, but necessary as no convenience method exists
    uncharged_parent_clean_mol = uncharger.uncharge(parent_clean_mol)

    # note that no attempt is made at reionization at this step
    # nor at ionization at some pH (rdkit has no pKa caculator)
    # the main aim to to represent all molecules from different sources
    # in a (single) standard way, for use in ML, catalogue, etc.

    te = rdMolStandardize.TautomerEnumerator()  # idem
    taut_uncharged_parent_clean_mol = te.Canonicalize(uncharged_parent_clean_mol)

    return taut_uncharged_parent_clean_mol


# adapted from plotnine package
def save_as_pdf_pages(
    plots,
    filename=None,
    path=None,
    verbose=True,
    **kwargs,
):
    # as in ggplot.save()
    fig_kwargs = {"bbox_inches": "tight"}
    fig_kwargs.update(kwargs)

    # If plots is already an iterator, this is a no-op; otherwise
    # convert a list, etc. to an iterator
    plots = iter(plots)

    # filename, depends on the object
    if filename is None:
        # Take the first element from the iterator, store it, and
        # use it to generate a file name
        peek = [next(plots)]
        plots = itertools.chain(peek, plots)
        filename = peek[0]._save_filename("pdf")

    if path:
        filename = pathlib.Path(path) / filename

    if verbose:
        warnings.warn(f"Filename: {filename}", p9.exceptions.PlotnineWarning)

    with PdfPages(filename, keep_empty=False) as pdf:
        # Re-add the first element to the iterator, if it was removed
        for plot in plots:
            if isinstance(plot, p9.ggplot):
                fig = plot.draw()
                with p9._utils.context.plot_context(plot).rc_context:
                    # Save as a page in the PDF file
                    pdf.savefig(fig, **fig_kwargs)
            elif isinstance(plot, plt.Figure) or isinstance(plot, matplotlib.table.Table):
                pdf.savefig(plot)
            else:
                raise TypeError(f"Unsupported type {type(plot)}. Must be ggplot or Figure.")


def parse_mgf_file(file_path):
    """
    Parses an MGF file and returns a dictionary containing the parsed data.

    Args:
        file_path (str): Path to the MGF file.

    Returns:
        dict: A dictionary where each key is a FEATURE_ID and the value is a list of blocks.
    """
    with open(file_path, "r", errors="ignore") as file:
        lines = file.readlines()

    blocks = []
    current_block = OrderedDict()

    required_keys = ["$$spectrumData", "pepmass", "instrument", "name", "adduct", "ionmode", "fragmentation_method", "collision_energy"]

    for line in lines:
        line = line.strip()
        if line == "BEGIN IONS":
            current_block = OrderedDict()

        elif line == "END IONS":
            missing_keys = [key for key in required_keys if key not in current_block.keys()]
            if len(missing_keys) > 0:
                print(f"\033[91mIncomplete block detected. Missing keys {str(missing_keys)}.\033[0m")
            blocks.append(current_block)
            current_block = OrderedDict()

        elif line == "":
            pass

        elif "=" in line:
            key, value = line.split("=", 1)
            key = key.strip().lower()
            value = value.strip()

            if key in ["collision_energy", "ms_ionization_energy"]:
                if re.match(r"^\d+(\.\d+)? *(HCD|CID)$", value, re.IGNORECASE):
                    match = re.match(r"^(\d+(\.\d+)?) *(HCD|CID)$", value, re.IGNORECASE)
                    if match:
                        value = match.group(1)
                        method_value = match.group(3).upper()
                        value = str(value)
                        current_block["fragmentation_method"] = method_value

                try:
                    if value.startswith("[") and value.endswith("]"):
                        # Already a list, normalize whitespace and sort
                        items = [float(v.strip()) for v in value[1:-1].split(",") if v.strip()]
                        items.sort()
                        value = items
                    elif "," in value:
                        # Comma-separated list, parse, sort, and format as python list string
                        items = [float(v.strip()) for v in value.split(",") if v.strip()]
                        items.sort()
                        value = items
                    else:
                        # Single value, wrap in list
                        try:
                            num = float(value)
                            value = [num]
                        except ValueError:
                            pass  # Leave as is if not a number
                except ValueError:
                    # If conversion fails, keep the original value
                    pass
                key = "collision_energy"

            elif key in ["fragmentation_mode", "ms_frag_mode", "fragmentation_method", "ms_dissociation_method"]:
                key = "fragmentation_method"

            elif key in ["ionmode", "ion_mode", "ms_ion_mode", "polarity"]:
                if key == "polarity":
                    if not value in ["1", "0"]:
                        raise ValueError(f"Invalid polarity value: {value}. Use '1' for positive and '0' for negative.")
                    value = "+" if value == "1" else "-"
                if value.lower() in ["positive", "pos", "p", "+"]:
                    value = "+"
                elif value.lower() in ["negative", "neg", "n", "-"]:
                    value = "-"
                else:
                    raise ValueError(f"Invalid ion mode: {value}")
                key = "ionmode"

            elif key in ["feature_id", "accession"]:
                key = "feature_id"

            elif key in ["precursor_type", "adduct"]:
                key = "adduct"

            elif key in ["precursormz", "precursor_mz", "pepmass", "precursor_mz_value"]:
                key = "pepmass"

            elif key in ["instrument", "instrument_model", "instrument_model_name", "instrument_name", "source_instrument", "ms_mass_analyzer"]:
                key = "instrument"

            elif key in ["name", "compound_name"]:
                value = value.lower().strip()
                key = "name"

            current_block[key] = str(value)

        elif line.lower().startswith("num peaks"):
            if "spectrumData" not in current_block:
                current_block["$$spectrumData"] = [[], []]

        else:
            if "$$spectrumData" not in current_block:
                current_block["$$spectrumData"] = [[], []]
            mz, inte = line.split()
            current_block["$$spectrumData"][0].append(mz)
            current_block["$$spectrumData"][1].append(inte)

    return blocks


def export_mgf_file(blocks, output_file_path):
    """
    Exports parsed MGF blocks to a new MGF file.

    Args:
        blocks (dict): Parsed MGF blocks.
        output_file_path (str): Path to the output MGF file.
    """
    with open(output_file_path, "w") as file:
        for feature_blocks in blocks:
            file.write("BEGIN IONS\n")
            for key, value in feature_blocks.items():
                if key == "$$spectrumData":
                    pass
                elif key.lower() == "collision_energy":
                    file.write(f"{key}={str(value).replace(' ', '')}\n")
                else:
                    file.write(f"{key}={value}\n")
            if "$$spectrumData" in feature_blocks:
                file.write("Num peaks={}\n".format(len(feature_blocks["$$spectrumData"][0])))
                for mzi in range(len(feature_blocks["$$spectrumData"][0])):
                    file.write(f"{feature_blocks['$$spectrumData'][0][mzi]} {feature_blocks['$$spectrumData'][1][mzi]}\n")
            file.write("END IONS\n\n")


def show_overview_of_blocks(blocks):
    """
    Prints an overview of the parsed MGF blocks, including counts of features, instruments, and fragmentation methods.
    Args:
        blocks (list): List of parsed MGF blocks.
    """

    fields = set()
    for block in blocks:
        fields.update(block.keys())

    print(f"   - {len(fields)} unique keys found in blocks: ", end="")
    for field in natsort.natsorted(fields, key=lambda x: x.lower()):
        print(f", {field}", end="")
    print("")

    instrument_counts = {}
    for block in blocks:
        if "instrument" in block:
            instrument = block["instrument"]
        else:
            instrument = "Unknown"
        instrument_counts[instrument] = instrument_counts.get(instrument, 0) + 1

    if instrument_counts:
        print("   - Instrument usage counts: ", end="")
        sorted_instruments = sorted(instrument_counts.items(), key=lambda item: item[1], reverse=True)
        for instrument, count in sorted_instruments:
            print(f", {instrument}: {count}", end="")
        print("")
    else:
        print("   - No instrument information found in blocks.")

    fragmentation_method_counts = {}
    for block in blocks:
        if "fragmentation_method" in block:
            method = block["fragmentation_method"]
        else:
            method = "Unknown"
        fragmentation_method_counts[method] = fragmentation_method_counts.get(method, 0) + 1

    if fragmentation_method_counts:
        print("   - Fragmentation method usage counts: ", end="")
        sorted_methods = sorted(fragmentation_method_counts.items(), key=lambda item: item[1], reverse=True)
        for method, count in sorted_methods:
            print(f", {method}: {count}", end="")
        print()
    else:
        print("   - No fragmentation method information found in blocks.")

    collision_energy_counts = {}
    for block in blocks:
        if "collision_energy" in block:
            energy = block["collision_energy"]
        else:
            energy = "Unknown"
        collision_energy_counts[energy] = collision_energy_counts.get(energy, 0) + 1

    if collision_energy_counts:
        print("   - Collision energy usage counts: ", end="")
        sorted_energies = sorted(collision_energy_counts.items(), key=lambda item: item[1], reverse=True)
        for energy, count in sorted_energies:
            print(f", {energy}: {count}", end="")
        print()
    else:
        print("   - No collision energy information found in blocks.")


def verify_mgf(blocks, smiles_field, name_field):
    """
    Verifies the integrity of MGF blocks by checking for required fields.

    Args:
        blocks (list): List of parsed MGF blocks.

    Returns:
        bool: True if all blocks are valid, False otherwise.
    """

    found = {}

    for block in blocks:
        smiles = None
        name = None

        if smiles_field in block:
            smiles = block[smiles_field]
            if smiles not in found:
                found[smiles] = set()

        if name_field in block:
            name = block[name_field]
            if name not in found:
                found[name] = set()

        if smiles is not None and name is not None:
            found[smiles].add(name.lower().strip())
            found[name].add(smiles.lower().strip())

    error = False
    for smiles, names in found.items():
        if len(names) > 1:
            print(f"Warning: name/smiles '{smiles}' is associated with multiple smiles/names: {', '.join(names)}")
            error = True

    return not error


def get_fields(blocks):
    """
    Extracts the unique fields from the parsed MGF blocks.    Returns:
        set: A set of unique field names found in the blocks.
    """
    fields = set()
    for block in blocks:
        fields.update(block.keys())
    return fields


def filter_blocks_with_required_fields(blocks):
    """
    Filters blocks to check if they contain the necessary fields 'pepmass' and 'rtinseconds'.

    Args:
        blocks (list): List of parsed MGF blocks.

    Returns:
        list: A list of blocks that contain both 'pepmass' and 'rtinseconds' fields.
    """
    filtered_blocks = []
    for block in blocks:
        keys = {key.lower() for key in block.keys()}
        if "pepmass" in keys and "rtinseconds" in keys:
            filtered_blocks.append(block)
    return filtered_blocks


def filter_low_intensity_peaks(blocks, intensity_threshold=0.01):
    """
    Filters out m/z and intensity pairs in the $$spectrumData field where the intensity
    is less than a specified percentage of the maximum intensity in the block.

    Args:
        blocks (list): List of parsed MGF blocks.
        intensity_threshold (float): The relative intensity threshold (default: 0.01).

    Returns:
        list: A list of blocks with filtered $$spectrumData.
    """
    for block in blocks:
        if "$$spectrumData" in block:
            mz, intensity = block["$$spectrumData"]
            max_intensity = np.max(intensity) if len(intensity) > 0 and np.sum(intensity) > 0 else 0
            if max_intensity > 0:
                use_inds = np.argwhere(intensity >= max_intensity * intensity_threshold).flatten()
                block["$$spectrumData"] = [mz[use_inds], intensity[use_inds]]
    return blocks


def filter_smiles(smarts_strings, check_fun):
    filtered_smiles = []
    non_matching_smiles = []
    errored_smiles = []

    for smiles in tqdm(smarts_strings):
        try:
            if check_fun(smiles):
                filtered_smiles.append(smiles)
            else:
                non_matching_smiles.append(smiles)
        except Exception as e:
            print(f"Error processing SMILES '{smiles}': {e}")
            errored_smiles.append(smiles)

    return filtered_smiles, non_matching_smiles, errored_smiles


def draw_smiles(smiles_strings, legends=None, max_draw=10, molsPerRow=10):
    if type(smiles_strings) is str or type(smiles_strings) is rdkit.Chem.rdchem.Mol:
        smiles_strings = [smiles_strings]

    mols = []
    for s in smiles_strings:
        add = []
        if type(s) is list or type(s) is rdkit.Chem.rdchem.Mol:
            add.extend(s)
        else:
            add = [s]

        for x in add:
            if type(x) == str:
                mol = rdkit.Chem.MolFromSmiles(x)
            elif type(x) is rdkit.Chem.rdchem.Mol:
                mol = x
            else:
                raise ValueError(f"Unsupported type for SMILES: {type(x)}")
            mols.append(mol)

    if len(smiles_strings) > max_draw:
        selected_indices = np.random.choice(len(smiles_strings), size=max_draw, replace=False)
        selected_mols = [mols[i] for i in selected_indices]
        selected_legends = [legends[i] for i in selected_indices] if legends else None
    else:
        selected_mols = mols
        selected_legends = legends if legends else None

    return rdkit.Chem.Draw.MolsToGridImage(
        selected_mols,
        legends=selected_legends,
        subImgSize=(500, 500),
        molsPerRow=min(molsPerRow, len(selected_mols)),
        maxMols=min(max_draw, len(selected_mols)),
        useSVG=False,
    )


def draw_names(compounds, blocks, name_field, smiles_field, max_draw=10, molsPerRow=10):
    """
    Plots the names of compounds as images.

    Args:
        cmpds (set): Set of compound names.
        blocks (list): List of parsed MGF blocks.
        max_draw (int): Maximum number of compounds to draw.
        molsPerRow (int): Number of molecules per row in the grid.
    """
    smiles = []
    legends = []
    for name in compounds:
        for block in blocks:
            if name_field in block.keys() and block[name_field] == name and smiles_field in block.keys() and block[smiles_field]:
                smiles.append(block[smiles_field])
                legends.append(name)
                break

    sorted_indices = sorted(range(len(legends)), key=lambda i: legends[i].lower())
    smiles = [smiles[i] for i in sorted_indices]
    legends = [legends[i] for i in sorted_indices]

    return draw_smiles(smiles, legends, max_draw, molsPerRow)


def substructure_fn(smiles, substructures_to_match=None):
    mol = rdkit.Chem.MolFromSmiles(smiles)

    if substructures_to_match is None:
        raise ValueError("No substructures to match provided")

    if type(substructures_to_match) is str or type(substructures_to_match) is rdkit.Chem.rdchem.Mol:
        substructures_to_match = [substructures_to_match]

    for sub in substructures_to_match:
        tomatch = None
        if type(sub) is list:
            tomatch = sub
        else:
            tomatch = [sub]

        anyMatch = False
        for smarts in tomatch:
            if type(smarts) is str:
                sub = rdkit.Chem.MolFromSmarts(smarts)
            elif type(smarts) is rdkit.Chem.rdchem.Mol:
                sub = smarts
            else:
                raise ValueError(f"Unsupported type for substructure: {type(smarts)}")
            anyMatch = anyMatch or mol.HasSubstructMatch(sub)

        if not anyMatch:
            return False
    return True


def prep_smarts_key(smart, replace=True):
    if replace:
        smart = smart.replace("c", "C").replace("o", "O").replace("C", "[C,c]").replace("O", "[O,o]")
    return rdkit.Chem.MolFromSmarts(smart)


def process_database(database_name, mgf_file, smiles_field, name_field, sf_field, smart_checks, output_folder, filter_fn=None, verbose=False):
    start = time.time()

    # parse MGF file
    spectra = parse_mgf_file(mgf_file)
    print(f"   - Parsed {len(spectra)} blocks")

    # filter for spectra that have smiles code
    spectra = [block for block in spectra if smiles_field in block.keys() and block[smiles_field] is not None and block[smiles_field].strip().lower() not in ["", "n/a", "na", "none", "null"]]
    print(f"   - Filtered to {len(spectra)} blocks with SMILES string, warning, must not be valid smiles")

    # show overview of the spectra
    show_overview_of_blocks(spectra)

    # apply custom filter if provided
    if filter_fn is not None:
        spectra = filter_fn(spectra)
        print(f"   - Filtered to {len(spectra)} blocks using custom filter")
        show_overview_of_blocks(spectra)

    # generate final data table for review
    table_data = OrderedDict()

    # get unique names and smiles from the spectra
    names = set([block[name_field] for block in spectra if name_field in block.keys() and block[name_field]])
    print(f"   - Found {len(names)} unique names in the MGF file")

    temp_dir = tempfile.TemporaryDirectory()

    # add each name to the table data
    for i, name in tqdm(enumerate(list(natsort.natsorted(names, key=lambda x: x.lower())))):
        # add compound
        table_data[name] = {"A_name": name, "A_comment": ""}

        # get unique smiles for the compound
        smiles = set([block[smiles_field] for block in spectra if name_field in block.keys() and block[name_field] == name and smiles_field in block.keys() and block[smiles_field]])
        table_data[name]["A_uniqueSmiles"] = str(smiles) if len(smiles) == 1 else f"{len(smiles)}: {smiles}"

        # get unique sum formulas
        formulas = set([block[sf_field] for block in spectra if name_field in block.keys() and block[name_field] == name and sf_field in block.keys() and block[sf_field]])
        table_data[name]["A_SumFormula"] = str(formulas) if len(formulas) == 1 else f"{len(formulas)}: {str(formulas)}"

        # draw structure if possible
        try:
            img = draw_smiles(smiles, max_draw=500)
            open(
                f"{temp_dir.name}/img_{i}.png",
                "wb",
            ).write(img.data)
            table_data[name]["A_structure"] = f"$$$IMG:{temp_dir.name}/img_{i}.png"
        except Exception as e:
            print(f"ERROR: could not draw structure for {name}: {e}")
            table_data[name]["A_structure"] = "ERROR: could not draw structure"

    n_spectra_with_smiles = sum(1 for block in spectra if smiles_field in block.keys() and block[smiles_field])
    print(f"   - Found {n_spectra_with_smiles} spectra with valid SMILES")

    unique_smiles_strings = sorted(list(set([block[smiles_field] for block in spectra if smiles_field in block.keys() and smiles_field != ""])))
    print(f"   - Found {len(unique_smiles_strings)} unique smiles strings")

    found_results = {}
    for check_name, check_parameters in smart_checks.items():
        found_results[check_name] = {}
        subs = check_parameters["filter"]

        print("\n\n--------------------------------------------------------------------------")
        print(f"   # Checking for substructure '{check_name}'")

        matching_smiles, non_matching_smiles, errored_smiles = filter_smiles(unique_smiles_strings, lambda x: substructure_fn(x, subs))
        found_results[check_name]["matching_smiles"] = matching_smiles

        if len(matching_smiles) > 0:
            matching_compounds = set()
            matching_blocks = []
            for spectrum in spectra:
                if smiles_field in spectrum.keys() and spectrum[smiles_field] in matching_smiles:
                    matching_compounds.add(spectrum[name_field])
                    matching_blocks.append(spectrum)

            print(f"   - Found {len(matching_compounds)} compounds with matching SMILES for '{check_name}'")
            for name in natsort.natsorted(matching_compounds, key=lambda x: x.lower()):
                table_data[name][f"C_{check_name}"] = "detected"
                if verbose:
                    print(f"      * {name}")
            found_results[check_name]["matching_compounds"] = matching_compounds

            if len(matching_smiles) > 0:
                try:
                    img = draw_names(matching_compounds, spectra, name_field, smiles_field, max_draw=500)
                    open(f"{output_folder}/{database_name}___{check_name}.png", "wb").write(img.data)
                except Exception as e:
                    print(f"ERROR: image generation failed, continuing without plotting substructures: {e}")

            write_path = f"{output_folder}/{database_name}___{check_name}.mgf"
            export_mgf_file(matching_blocks, write_path)
            print(f"   - Exported spectra to {write_path}")

        else:
            print("   - No matches found")

        print("")

    list_to_excel_table([v for k, v in table_data.items()], f"{output_folder}/{database_name}___table.xlsx", column_width=40, row_height=40)
    temp_dir.cleanup()

    print(f"   - Exported table to {output_folder}/{database_name}___table.xlsx")
    end = time.time()
    print(f"   - Processed {len(spectra)} spectra in {end - start:.2f} seconds")
    print("--------------------------------------------------------------------------\n")

    return found_results, spectra
