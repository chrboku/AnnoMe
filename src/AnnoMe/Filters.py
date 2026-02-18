import json
import os
import pathlib
import tempfile
import warnings
import re
import itertools
from collections import OrderedDict, defaultdict
import time
import requests
import csv

import numpy as np

import matplotlib
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import plotnine as p9
import polars as pl
import pandas as pd

import rdkit
import rdkit.Chem
import rdkit.Chem.Draw
from rdkit.Chem.Draw import IPythonConsole

# from rdkit import Chem
# from rdkit.Chem import Draw
# from rdkit.Chem.MolStandardize import rdMolStandardize

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from tqdm import tqdm

import natsort

import colorama
from colorama import Fore, Style

import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed


def optimize_dataframe_dtypes(df, infer_schema_length=0, max_values_for_categorical_column=1000):
    """
    Optimize DataFrame column datatypes to reduce memory usage.
    Supports both Polars and pandas DataFrames.

    This function analyzes each column and attempts to convert it to the most memory-efficient
    datatype while preserving data integrity:
    1. Try converting to Int64 (most memory-efficient for integers)
    2. If that fails, try Float64 (for numeric data with decimals)
    3. If that fails, keep as string
    4. For string columns with few unique values, convert to Categorical (more memory-efficient)

    Args:
        df: Polars or pandas DataFrame to optimize
        infer_schema_length: Number of non-empty rows to sample per column for type inference.
            If <= 0, uses all rows (default: 0)
        max_values_for_categorical_column: Maximum number of unique values to convert
            a string column to categorical (default: 1000)

    Returns:
        Optimized DataFrame with more memory-efficient dtypes (same type as input)
    """

    if isinstance(df, pl.DataFrame):
        return _optimize_polars_dtypes(df, infer_schema_length, max_values_for_categorical_column)
    elif isinstance(df, pd.DataFrame):
        return _optimize_pandas_dtypes(df, infer_schema_length, max_values_for_categorical_column)
    else:
        raise TypeError(f"Unsupported DataFrame type: {type(df)}. Expected polars.DataFrame or pandas.DataFrame.")


def _optimize_polars_dtypes(df, infer_schema_length=0, max_values_for_categorical_column=1000):
    """
    Optimize Polars DataFrame column datatypes to reduce memory usage.

    Args:
        df: Polars DataFrame to optimize
        infer_schema_length: Number of non-empty rows to sample per column for type inference.
            If <= 0, uses all rows
        max_values_for_categorical_column: Maximum number of unique values to convert
            a string column to categorical

    Returns:
        Optimized Polars DataFrame with more memory-efficient dtypes
    """
    if df is None or df.is_empty():
        return df

    estimated_size_before = df.estimated_size()

    if infer_schema_length <= 0:
        print(f"Optimizing Polars DataFrame dtypes (using all rows)...")
    else:
        print(f"Optimizing Polars DataFrame dtypes (sampling {infer_schema_length} non-empty rows per column)...")

    num_conversions = 0

    for col_name in df.columns:
        # Only optimize string-like columns
        if df[col_name].dtype != pl.Utf8:
            continue

        # Get non-null values for this column
        non_null_sample = df.select(pl.col(col_name)).filter(pl.col(col_name).is_not_null())

        # Limit to infer_schema_length rows if specified (> 0)
        if infer_schema_length > 0:
            non_null_sample = non_null_sample.head(infer_schema_length)

        if non_null_sample.is_empty():
            # All nulls, skip optimization
            continue

        # Try converting to Int64
        try:
            non_null_sample.select(pl.col(col_name).cast(pl.Int64, strict=True))
            # Success! Convert entire column to Int64 immediately
            df = df.with_columns(pl.col(col_name).cast(pl.Int64, strict=False))
            print(f"  - {col_name}: Utf8 → Int64")
            num_conversions += 1
            continue
        except:
            pass

        # Try converting to Float64
        try:
            non_null_sample.select(pl.col(col_name).cast(pl.Float64, strict=True))
            # Success! Convert entire column to Float64 immediately
            df = df.with_columns(pl.col(col_name).cast(pl.Float64, strict=False))
            print(f"  - {col_name}: Utf8 → Float64")
            num_conversions += 1
            continue
        except:
            pass

        # Keep as Utf8, but check if it should be Categorical
        # Count unique values in the sample
        unique_count = non_null_sample.select(pl.col(col_name).n_unique()).item()

        if unique_count <= max_values_for_categorical_column:
            # Convert to Categorical for memory efficiency immediately
            df = df.with_columns(pl.col(col_name).cast(pl.Categorical))
            print(f"  - {col_name}: Utf8 → Categorical ({unique_count} unique values)")
            num_conversions += 1

    # Shrink to fit after all conversions
    if num_conversions > 0:
        df = df.shrink_to_fit()
        print(
            f"Optimized {num_conversions} column(s), reduced size by {(1.0 - df.estimated_size() / estimated_size_before) * 100.0:.2f}% (size before: {estimated_size_before} bytes, size after: {df.estimated_size()} bytes)"
        )
    else:
        print("No columns required optimization")

    return df


def _optimize_pandas_dtypes(df, infer_schema_length=0, max_values_for_categorical_column=1000):
    """
    Optimize pandas DataFrame column datatypes to reduce memory usage.

    Args:
        df: pandas DataFrame to optimize
        infer_schema_length: Number of non-empty rows to sample per column for type inference.
            If <= 0, uses all rows
        max_values_for_categorical_column: Maximum number of unique values to convert
            a string column to categorical

    Returns:
        Optimized pandas DataFrame with more memory-efficient dtypes
    """
    import pandas as pd

    if df is None or df.empty:
        return df

    estimated_size_before = df.memory_usage(deep=True).sum()

    if infer_schema_length <= 0:
        print(f"Optimizing pandas DataFrame dtypes (using all rows)...")
    else:
        print(f"Optimizing pandas DataFrame dtypes (sampling {infer_schema_length} non-empty rows per column)...")

    num_conversions = 0

    for col_name in df.columns:
        # Only optimize object (string) columns
        if df[col_name].dtype != object:
            continue

        # Get non-null values for this column
        non_null_series = df[col_name].dropna()

        # Limit to infer_schema_length rows if specified (> 0)
        if infer_schema_length > 0:
            non_null_series = non_null_series.head(infer_schema_length)

        if non_null_series.empty:
            # All nulls, skip optimization
            continue

        # Try converting to Int64 (nullable integer)
        try:
            converted = pd.to_numeric(non_null_series, errors="raise")
            # Check if all values are integers (no fractional parts)
            if (converted == converted.astype("int64")).all():
                df[col_name] = pd.to_numeric(df[col_name], errors="coerce").astype("Int64")
                print(f"  - {col_name}: object → Int64")
                num_conversions += 1
                continue
        except (ValueError, TypeError, OverflowError):
            pass

        # Try converting to Float64
        try:
            pd.to_numeric(non_null_series, errors="raise")
            df[col_name] = pd.to_numeric(df[col_name], errors="coerce").astype("float64")
            print(f"  - {col_name}: object → float64")
            num_conversions += 1
            continue
        except (ValueError, TypeError):
            pass

        # Keep as object, but check if it should be Categorical
        try:
            unique_count = non_null_series.nunique()

            if unique_count <= max_values_for_categorical_column:
                df[col_name] = df[col_name].astype("category")
                print(f"  - {col_name}: object → category ({unique_count} unique values)")
                num_conversions += 1
                continue
        except Exception as e:
            pass

    if num_conversions > 0:
        estimated_size_after = df.memory_usage(deep=True).sum()
        print(
            f"Optimized {num_conversions} column(s), reduced size by {(1.0 - estimated_size_after / estimated_size_before) * 100.0:.2f}% (size before: {estimated_size_before} bytes, size after: {estimated_size_after} bytes)"
        )
    else:
        print("No columns required optimization")

    return df


def download_file_if_not_exists(url, dest_folder, file_name=None, print_intention=0, status_bar_update_func=None, status_bar_max_func=None):
    """
    Downloads a file from a given URL to a specified destination folder.

    Args:
        url (str): The URL of the file to download.
        dest_folder (str): The folder where the downloaded file will be saved.
        file_name (str, optional): The name to give the downloaded file. If not provided, the original file name will be used.

    Returns:
        str: The path to the downloaded file.
    """
    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)

    if file_name is None:
        filename = os.path.join(dest_folder, url.split("/")[-1])
    else:
        filename = os.path.join(dest_folder, file_name)

    try:
        response = requests.get(url, stream=True)

        # show a ajax loader animation with an update every second
        total_size = int(response.headers.get("content-length", 0))

        # check if the file is present in this size
        if os.path.exists(filename):
            existing_size = os.path.getsize(filename)
            if existing_size == total_size:
                print(f"{' ' * print_intention}- File already downloaded.")
                return filename
            else:
                print(f"{' ' * print_intention}{Fore.RED}- Partial file found, re-downloading...{Style.RESET_ALL}")
                os.remove(filename)

        if total_size > 100_000_000:  # > 100MB
            print_size = total_size / (1000 * 1000)
            print_size_unit = "MB"
            if print_size > 1000:
                print_size = print_size / 1000
                print_size_unit = "GB"
            print(f"{' ' * print_intention}{Fore.YELLOW}Warning: File is large ({print_size:.2f} {print_size_unit}), this may take a while...{Style.RESET_ALL}")

        block_size = 1000  # 1 Kibibyte

        t = tqdm(total=total_size, unit="iB", unit_scale=True)
        if status_bar_update_func is not None:
            status_bar_max_func(total_size)

        total_size_loaded = 0
        with open(filename, "wb") as f:
            for data in response.iter_content(block_size):
                f.write(data)

                total_size_loaded += len(data)
                t.update(len(data))
                if status_bar_update_func is not None:
                    status_bar_update_func(total_size_loaded)

    except Exception as e:
        print(f"{' ' * print_intention}- Error downloading {url}: {e}")
        try:
            os.remove(filename)
        except Exception as e:
            print(f"{' ' * print_intention}- Error removing {filename}: {e}")
        raise RuntimeError(f"Failed to download {url} to {filename}")

    return filename


def tsv_to_mgf(tsv_file_path, mgf_file_path):
    """
    Converts a TSV file to an MGF file. Each row becomes a block, with column names as keys.

    Args:
        tsv_file_path (str): Path to the input TSV file.
        mgf_file_path (str): Path to the output MGF file.
    """

    with (
        open(tsv_file_path, "r", encoding="utf-8") as tsvfile,
        open(mgf_file_path, "w", encoding="utf-8") as mgffile,
    ):
        reader = csv.DictReader(tsvfile, delimiter="\t")
        for row in reader:
            mgffile.write("BEGIN IONS\n")
            for key, value in row.items():
                if key in ["mzs", "intensities"]:
                    continue
                if key == "identifier":
                    key = "name"
                if key == "precursor_mz":
                    key = "pepmass"
                if value is not None and value != "":
                    mgffile.write(f"{key}={value}\n")
            for mz, inte in zip(row["mzs"].split(","), row["intensities"].split(",")):
                if mz.strip() and inte.strip():
                    mgffile.write(f"{mz.strip()} {inte.strip()}\n")
            mgffile.write("END IONS\n\n")


def msp_to_mgf(msp_file_path, mgf_file_path):
    """
    Converts an MSP file to an MGF file.
    Each MSP block is converted to an MGF block, with key-value pairs using '=' and no 'Num Peaks' line.

    Args:
        msp_file_path (str): Path to the input MSP file.
        mgf_file_path (str): Path to the output MGF file.
    """

    def parse_block(block_lines):
        header = []
        peaks = []
        for line in block_lines:
            if not line.strip():
                continue
            if ":" in line:
                key, value = line.split(":", 1)
                key = key.strip()
                value = value.strip()
                if key.lower() == "num peaks":
                    continue
                header.append((key, value))
            else:
                # Assume it's a peak line (mz intensity)
                if "\t" in line:
                    mz, intensity = line.split("\t", 1)
                elif " " in line:
                    mz, intensity = line.split(None, 1)
                else:
                    continue
                peaks.append((mz.strip(), intensity.strip()))
        return header, peaks

    with (
        open(msp_file_path, "r", encoding="utf-8") as infile,
        open(mgf_file_path, "w", encoding="utf-8") as outfile,
    ):
        block_lines = []
        for line in infile:
            if line.strip() == "" and block_lines:
                header, peaks = parse_block(block_lines)
                outfile.write("BEGIN IONS\n")
                for key, value in header:
                    if key.lower() == "PRECURSORMZ":
                        key = "pepmass"
                    if key.lower() == "PRECURSORTYPE":
                        key = "adduct"
                    if key.lower() == "RETENTIONTIME":
                        key = "rtinseconds"
                    if key.lower() == "PRECURSORMZ":
                        key = "pepmass"
                    if key.lower() == "PRECURSORMZ":
                        key = "pepmass"
                    outfile.write(f"{key}={value}\n")
                for mz, intensity in peaks:
                    outfile.write(f"{mz} {intensity}\n")
                outfile.write("END IONS\n\n")
                block_lines = []
            else:
                block_lines.append(line)
        # Write last block if file doesn't end with newline
        if block_lines:
            header, peaks = parse_block(block_lines)
            outfile.write("BEGIN IONS\n")
            for key, value in header:
                outfile.write(f"{key}={value}\n")
            for mz, intensity in peaks:
                outfile.write(f"{mz} {intensity}\n")
            outfile.write("END IONS\n\n")


def fix_massspecgym_nameandid(mgf_file_path):
    """
    Fixes the 'name' and 'id' fields in the MassSpecGym MGF file.
    The 'name' field is set to the 'id' field, and the 'id' field is removed.

    Args:
        mgf_file_path (str): Path to the MassSpecGym MGF file.
    """
    with open(mgf_file_path, "r", encoding="utf-8") as infile:
        lines = infile.readlines()

    with open(mgf_file_path, "w", encoding="utf-8") as outfile:
        for line in lines:
            if line.startswith("name="):
                # Set name to id
                outfile.write(line.replace("name=", "massspecgym_id="))

            elif line.startswith("inchikey="):
                # Set id to name
                outfile.write(line)
                outfile.write(line.replace("inchikey=", "name="))

            else:
                outfile.write(line)


def get_config_value(key, config_location=None):
    """
    Retrieves a configuration value from a specified location or the default config file.

    Args:
        key (str or list of str): The key(s) for which to retrieve the value.
            - If str: Returns the value from the top-level dictionary.
            - If list of str: Navigates through nested levels and returns the final value.
        config_location (str, optional): The path to the configuration file. If None, uses the default location.

    Returns:
        The value associated with the key, or None if not found.

    Examples:
        get_config_value("MONA") -> returns the MONA dictionary
        get_config_value(["MONA", "url"]) -> returns the URL within the MONA dictionary
        get_config_value(["MSnLib_files", "1", "file_name"]) -> returns the file_name in entry "1"
    """
    if config_location is None:
        config_location = os.path.join(os.path.dirname(__file__), "../../config.json")

    if not os.path.exists(config_location):
        raise FileNotFoundError(f"Configuration file not found at {config_location}")

    with open(config_location, "r") as f:
        config = json.load(f)

        # Handle single string key (top-level access)
        if isinstance(key, str):
            if key not in config:
                raise KeyError(f"Key '{key}' not found in configuration file at {config_location}")
            return config.get(key, None)

        # Handle list of keys (nested access)
        elif isinstance(key, list):
            current = config
            for i, k in enumerate(key):
                if not isinstance(current, dict):
                    raise TypeError(f"Cannot access key '{k}' at level {i}: current value is not a dictionary")
                if k not in current:
                    key_path = " -> ".join(key[: i + 1])
                    raise KeyError(f"Key path '{key_path}' not found in configuration file at {config_location}")
                current = current[k]
            return current

        else:
            raise TypeError(f"Key must be a string or list of strings, got {type(key)}")

    raise RuntimeError(f"Failed to retrieve value for key '{key}' from configuration file at {config_location}, unknown error")


def get_config_keys(key=None, config_location=None):
    """
    Returns all keys at a particular level in the configuration file.

    Args:
        key (str, list of str, or None): The key(s) specifying which level to query.
            - If None: Returns all keys from the top-level dictionary.
            - If str: Returns all keys within the specified top-level dictionary.
            - If list of str: Navigates through nested levels and returns keys at that level.
        config_location (str, optional): The path to the configuration file. If None, uses the default location.

    Returns:
        list: A list of keys at the specified level.

    Examples:
        get_config_keys() -> returns all top-level keys ["MONA", "GNPS_Wine_DB_Orbitrap", ...]
        get_config_keys("MONA") -> returns keys within MONA ["url", "size"]
        get_config_keys(["MSnLib_files"]) -> returns keys within MSnLib_files ["base_url", "1", "2", ...]
        get_config_keys(["MSnLib_files", "1"]) -> returns keys within entry "1" ["file_name", "size"]
    """
    if config_location is None:
        config_location = os.path.join(os.path.dirname(__file__), "../../config.json")

    if not os.path.exists(config_location):
        raise FileNotFoundError(f"Configuration file not found at {config_location}")

    with open(config_location, "r") as f:
        config = json.load(f)

        # Handle None (return top-level keys)
        if key is None:
            return list(config.keys())

        # Handle single string key
        if isinstance(key, str):
            if key not in config:
                raise KeyError(f"Key '{key}' not found in configuration file at {config_location}")
            value = config[key]
            if not isinstance(value, dict):
                raise TypeError(f"Value at key '{key}' is not a dictionary, cannot retrieve keys")
            return list(value.keys())

        # Handle list of keys (nested access)
        elif isinstance(key, list):
            current = config
            for i, k in enumerate(key):
                if not isinstance(current, dict):
                    raise TypeError(f"Cannot access key '{k}' at level {i}: current value is not a dictionary")
                if k not in current:
                    key_path = " -> ".join(key[: i + 1])
                    raise KeyError(f"Key path '{key_path}' not found in configuration file at {config_location}")
                current = current[k]

            if not isinstance(current, dict):
                key_path = " -> ".join(key)
                raise TypeError(f"Value at key path '{key_path}' is not a dictionary, cannot retrieve keys")
            return list(current.keys())

        else:
            raise TypeError(f"Key must be None, a string, or list of strings, got {type(key)}")


def unzip_file(zip_file_path, dest_folder):
    """
    Unzips a zip file to the specified destination folder.

    Args:
        zip_file_path (str): The path to the zip file.
        dest_folder (str): The folder where the contents will be extracted.
    """

    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)

    with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
        zip_ref.extractall(dest_folder)


def download_MSMS_libraries(dest_folder=None, status_bar_update_func=None, status_bar_max_func=None, status_bar_description_func=None):
    """
    Downloads common MS/MS libraries to the specified destination folder.

    Args:
        dest_folder (str): The folder where the libraries will be downloaded.
    """

    if dest_folder is None:
        dest_folder = os.path.join(".", "resources")

    c_dest_folder = os.path.join(dest_folder, "libraries")

    print(f"Downloading (if necessary) public datasets for MS/MS libraries, to {c_dest_folder}...")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading MS/MS libraries...")

    # Mona
    print(f"{Fore.GREEN}   - MONA{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading MONA library...")
    download_file_if_not_exists(get_config_value(["MONA", "url"]), c_dest_folder, print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func)

    # GNPS datasets
    print(f"{Fore.GREEN}   - GNPS Wine DB Orbitrap{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading GNPS Wine DB Orbitrap library...")
    download_file_if_not_exists(
        get_config_value(["GNPS_Wine_DB_Orbitrap", "url"]), c_dest_folder, print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func
    )
    print(f"{Fore.GREEN}   - GNPS cleaned{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading GNPS cleaned library...")
    download_file_if_not_exists(get_config_value(["GNPS_Cleaned", "url"]), c_dest_folder, print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func)

    # MassSpecGym and conversion thereof to mgf file
    print(f"{Fore.GREEN}   - MassSpecGym{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading MassSpecGym library...")
    download_file_if_not_exists(get_config_value(["MassSpecGym", "url"]), c_dest_folder, print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func)
    if os.path.exists(os.path.join(c_dest_folder, "MassSpecGym.mgf")):
        print("      - File already processeded")
    else:
        print("      - processing")
        tsv_to_mgf(
            os.path.join(c_dest_folder, "MassSpecGym.tsv"),
            os.path.join(c_dest_folder, "MassSpecGym.mgf"),
        )
        fix_massspecgym_nameandid(os.path.join(c_dest_folder, "MassSpecGym.mgf"))

    # MassBank Riken and conversion thereof to mgf file
    print(f"{Fore.GREEN}   - MassBank Riken{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading MassBank Riken library...")
    download_file_if_not_exists(get_config_value(["MB_Riken", "url"]), c_dest_folder, print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func)
    if os.path.exists(os.path.join(c_dest_folder, "MassBank_RIKEN.mgf")):
        print("      - File already processeded")
    else:
        print("      - processing")
        msp_to_mgf(
            os.path.join(c_dest_folder, "MassBank.msp_RIKEN"),
            os.path.join(c_dest_folder, "MassBank_RIKEN.mgf"),
        )

    # MSnLib
    base_url = get_config_value(["MSnLib_files", "base_url"])
    files = get_config_value("MSnLib_files")
    print(f"{Fore.GREEN}   - MSnLib files:{Style.RESET_ALL}")
    for file_name in files:
        if file_name == "base_url":
            continue
        print(f"{Fore.GREEN}      - {file_name}{Style.RESET_ALL}")
        if status_bar_description_func is not None:
            status_bar_description_func(f"Downloading MSnLib file {file_name}...")
        download_file_if_not_exists(
            base_url.format(file_name=f"{file_name}.mgf"),
            c_dest_folder,
            file_name=f"{file_name}.mgf",
            print_intention=9,
            status_bar_update_func=status_bar_update_func,
            status_bar_max_func=status_bar_max_func,
        )

    # BOKU iBAM in-house database (more specific towards example dataset)
    print(f"{Fore.GREEN}   - BOKU iBAM{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading BOKU iBAM library...")
    download_file_if_not_exists(
        get_config_value(["BOKU_iBAM", "url"]), c_dest_folder, file_name="BOKU_iBAM.mgf", print_intention=6, status_bar_update_func=status_bar_update_func, status_bar_max_func=status_bar_max_func
    )

    c_dest_folder = os.path.join(dest_folder, "libraries_other")

    # other libraries
    print(f"{Fore.GREEN}   - other MSMS datasets{Style.RESET_ALL}")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading other MSMS datasets...")
    download_file_if_not_exists(
        get_config_value(["BOKU_other_MSMS_datasets", "url"]),
        c_dest_folder,
        file_name="other_MSMS_datasets.zip",
        print_intention=6,
        status_bar_update_func=status_bar_update_func,
        status_bar_max_func=status_bar_max_func,
    )
    # Unzip the downloaded file if it is a zip file
    unzip_file(os.path.join(c_dest_folder, "other_MSMS_datasets.zip"), c_dest_folder)


def download_MS2DeepScore_model(dest_folder=None, status_bar_update_func=None, status_bar_max_func=None, status_bar_description_func=None):
    """
    Downloads the MS2DeepScore model to the specified destination folder.
    """

    if dest_folder is None:
        dest_folder = os.path.join(".", "resources", "models")

    print(f"Downloading (if necessary) {Fore.GREEN}MS2DeepScore model{Style.RESET_ALL}, to {dest_folder}...")
    if status_bar_description_func is not None:
        status_bar_description_func("Downloading MS2DeepScore model...")
    download_file_if_not_exists(
        get_config_value(["MS2DeepScore_model", "url"]),
        dest_folder,
        file_name="ms2deepscore_model.pt",
        print_intention=3,
        status_bar_update_func=status_bar_update_func,
        status_bar_max_func=status_bar_max_func,
    )


def CE_parser(ce_str):
    """
    Parses the collision energy string to extract the numeric value.
    Handles both absolute and relative values.
    """
    if isinstance(ce_str, str):
        # Try to match patterns like '70 eV (absolute)', '75eV', '75.0eV', '15V'
        match = re.search(r"(\d+(?:\.\d+)?)\s*(eV|V)?", ce_str, re.IGNORECASE)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                pass

        # Try to match patterns like '[45.0]', '[45]'
        match = re.match(r"\[(\d+(?:\.\d+)?)\]", ce_str)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                pass
    return -1


def scale_dimensions_to_fit(original_width, original_height, max_width, max_height):
    """
    Scales (width, height) to fit within max dimensions, preserving aspect ratio.

    Returns: (new_width, new_height)
    """

    asp_ratio = original_width / original_height

    new_width = max_width
    new_height = max_width / asp_ratio

    if new_height > max_height:
        new_height = max_height
        new_width = max_height * asp_ratio

    return int(new_width), int(new_height)


def write_to_excel_cell(
    filename,
    sheet_name,
    texts,
    images=None,
    column_width=None,
    row_height=None,
    img_scale_fact=1.33,
):
    """
    Writes text and images to a specific Excel sheet.
    Args:
        filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to write to.
        texts (dict): A dictionary mapping cell addresses (e.g., 'A1') to text content.
        images (dict, optional): A dictionary mapping cell addresses to image file paths.
        column_width (float, optional): Width for all columns.
        row_height (float, optional): Height for all rows.
        img_scale_fact (float, optional): Scale factor for images. Defaults to 1.33.
    """
    # Create a new workbook if file doesn't exist, else load it workbook if file doesn't exist, else load it
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    # Get or create the desired sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    for cell, content in texts.items():
        ws[cell] = content
        col_letter = "".join(filter(str.isalpha, cell))
        row_number = int("".join(filter(str.isdigit, cell)))
        if column_width is not None:
            ws.column_dimensions[col_letter].width = column_width
        if row_height is not None:
            ws.row_dimensions[row_number].height = row_height

    if images is not None:
        for cell, image in images.items():
            if not os.path.exists(image):
                raise FileNotFoundError(f"Image file '{image}' not found.")
            # Load image and add to sheetimg_scale_factimg_scale_fact
            img = ExcelImage(image)
            img.width, img.height = scale_dimensions_to_fit(
                img.width,
                img.height,
                column_width * img_scale_fact,
                row_height * img_scale_fact,
            )

            # Position image to the top-left of the cell
            ws.add_image(img, cell)

            # Optional: Adjust column width and row height
            col_letter = "".join(filter(str.isalpha, cell))
            row_number = int("".join(filter(str.isdigit, cell)))
            ws.column_dimensions[col_letter].width = column_width
            ws.row_dimensions[row_number].height = row_height

    wb.save(filename)


def write_table_to_excel(
    filename,
    table,
    sheet_name="Sheet",
    img_prefix="$$$IMG:",
    column_width=40,
    row_height=8,
):
    """
    Writes a table (list of lists) to an Excel file using the write_to_excel_cell function.
    Deletes the file first if it exists.

    Args:
        filename (str): The path to the Excel file.
        table (list of list): The data to write, where each inner list is a row.
        sheet_name (str, optional): The name of the sheet to write to. Defaults to "Sheet1".

    Raises:
        OSError: If the existing file cannot be deleted.
    """
    if os.path.exists(filename):
        try:
            os.remove(filename)
        except OSError as e:
            raise OSError(f"Error: Could not delete existing file '{filename}'. Reason: {e}")

    texts_to_write = {}
    images_to_write = {}
    for r_idx, row_data in enumerate(table):
        for c_idx, cell_value in enumerate(row_data):
            # Convert 0-based c_idx to 1-based column letter
            col_letter = get_column_letter(c_idx + 1)
            # Excel rows are 1-based
            cell_address = f"{col_letter}{r_idx + 1}"

            if isinstance(cell_value, str) and cell_value.startswith(img_prefix):
                image_path = cell_value.replace(img_prefix, "", 1)
                # Ensure 'images_to_write' is initialized in the function scope, e.g., images_to_write = {}
                images_to_write[cell_address] = image_path
                # Set cell_value to an empty string so the subsequent line doesn't write the image directive as text
            else:
                # If it's not an image string, cell_value remains unchanged,
                # and the original `texts_to_write[cell_address] = cell_value` line will handle it as text.
                texts_to_write[cell_address] = cell_value

    # It will create a new workbook since we've deleted the file if it existed.
    write_to_excel_cell(
        filename,
        sheet_name,
        texts_to_write,
        images=images_to_write,
        column_width=column_width,
        row_height=row_height,
    )


def list_to_excel_table(
    data,
    filename,
    sheet_name="Sheet",
    img_prefix="$$$IMG:",
    column_width=40,
    row_height=8,
):
    """
    Converts a list of dictionaries into a table and writes it to an Excel file.

    The input list contains the rows and their data. The keys of these
    inner dictionaries are used as column headers.

    Args:
        data_dict (list): A list of dictionaries mapping column headers (strings)
                            to cell values.
                            Example: {"Row1": {"ColA": 1, "ColB": "img_path.png"}, "Row2": {"ColA": 2}}
        filename (str): The path to the Excel file to be created/overwritten.
        sheet_name (str, optional): The name of the sheet to write to. Defaults to "Sheet".
        row_key_col_name (str, optional): The header for the first column, which will
                                            contain the row keys from data_dict. Defaults to "Row_Key".
        img_prefix (str, optional): The prefix used to identify image paths in cell values.
                                    Defaults to "$$$IMG:".
    """
    if not isinstance(data, list):
        raise TypeError("Input datalist must be a list.")

    # Extract all unique keys from the inner dictionaries to form column headers
    unique_column_headers = set()
    if data:
        for inner_dict in data:
            if isinstance(inner_dict, dict):
                unique_column_headers.update(inner_dict.keys())
            # If inner_dict is not a dict, its keys won't be added.
            # This means columns are only derived from valid inner dictionaries.
        all_column_headers = natsort.natsorted(list(unique_column_headers))

        # Prepare the table (list of lists)
        table_data = []

        # Add the header row
        header_row = all_column_headers
        table_data.append(header_row)

        # Add the data rows
        for inner_dict in data:
            current_row_values = []
            if isinstance(inner_dict, dict):
                for col_header in all_column_headers:
                    current_row_values.append(inner_dict.get(col_header, ""))  # Use empty string for missing values
            else:
                # If the value for a row_key is not a dict, fill data cells with empty strings
                current_row_values.extend([""] * len(all_column_headers))
            table_data.append(current_row_values)

        # Write the table to Excel using the provided function
        write_table_to_excel(
            filename,
            table_data,
            sheet_name=sheet_name,
            img_prefix=img_prefix,
            column_width=column_width,
            row_height=row_height,
        )


def standardize_smile(smiles):
    """
    Standardizes a SMILES string by removing hydrogens, disconnecting metal atoms,
    normalizing the molecule, and reionizing it. It also handles tautomer enumeration.
    Args:
        smiles (str): The SMILES string to standardize.
    """
    # follows the steps in
    # https://github.com/greglandrum/RSC_OpenScience_Standardization_202104/blob/main/MolStandardize%20pieces.ipynb
    # as described **excellently** (by Greg) in
    # https://www.youtube.com/watch?v=eWTApNX8dJQ
    mol = rdkit.Chem.MolFromSmiles(smiles)

    # removeHs, disconnect metal atoms, normalize the molecule, reionize the molecule
    clean_mol = rdMolStandardize.Cleanup(mol)

    # if many fragments, get the "parent" (the actual mol we are interested in)
    parent_clean_mol = rdMolStandardize.FragmentParent(clean_mol)

    # try to neutralize molecule
    uncharger = rdMolStandardize.Uncharger()  # annoying, but necessary as no convenience method exists
    uncharged_parent_clean_mol = uncharger.uncharge(parent_clean_mol)

    # note that no attempt is made at reionization at this step
    # nor at ionization at some pH (rdkit has no pKa caculator)
    # the main aim to to represent all molecules from different sources
    # in a (single) standard way, for use in ML, catalogue, etc.

    te = rdMolStandardize.TautomerEnumerator()  # idem
    taut_uncharged_parent_clean_mol = te.Canonicalize(uncharged_parent_clean_mol)

    return taut_uncharged_parent_clean_mol


# adapted from plotnine package
def save_as_pdf_pages(
    plots,
    filename=None,
    path=None,
    verbose=True,
    **kwargs,
):
    """
    Save multiple ggplot objects or matplotlib figures to a single PDF file.
    Args:
        plots (iterable): An iterable of ggplot objects or matplotlib figures.
        filename (str, optional): The name of the output PDF file. If not provided,
                                  it will be generated from the first plot.
        path (str, optional): The directory where the PDF file will be saved.
                              If not provided, the current working directory is used.
        verbose (bool, optional): If True, print the filename being saved.
                                  Defaults to True.
        **kwargs: Additional keyword arguments passed to PdfPages.savefig().
    """
    # as in ggplot.save()
    fig_kwargs = {"bbox_inches": "tight"}
    fig_kwargs.update(kwargs)

    # If plots is already an iterator, this is a no-op; otherwise
    # convert a list, etc. to an iterator
    plots = iter(plots)

    # filename, depends on the object
    if filename is None:
        # Take the first element from the iterator, store it, and
        # use it to generate a file name
        peek = [next(plots)]
        plots = itertools.chain(peek, plots)
        filename = peek[0]._save_filename("pdf")

    if path:
        filename = pathlib.Path(path) / filename

    if verbose:
        warnings.warn(f"Filename: {filename}", p9.exceptions.PlotnineWarning)

    with PdfPages(filename, keep_empty=False) as pdf:
        # Re-add the first element to the iterator, if it was removed
        for plot in plots:
            if isinstance(plot, p9.ggplot):
                fig = plot.draw()
                with p9._utils.context.plot_context(plot).rc_context:
                    # Save as a page in the PDF file
                    pdf.savefig(fig, **fig_kwargs)
            elif isinstance(plot, plt.Figure) or isinstance(plot, matplotlib.table.Table):
                pdf.savefig(plot)
            else:
                raise TypeError(f"Unsupported type {type(plot)}. Must be ggplot or Figure.")


def is_float(value):
    """
    Checks if a value can be converted to a float.

    Args:
        value: The value to check.

    Returns:
        bool: True if the value can be converted to a float, False otherwise.
    """
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def parse_mgf_file(file_path, check_required_keys=True, return_as_polars_table=False, ignore_spectral_data=False):
    """
    Parses an MGF file and returns a dictionary containing the parsed data.

    Args:
        file_path (str): Path to the MGF file.
        check_required_keys (bool): If True, only blocks with 'pepmass' and 'name' are kept.
        return_as_polars_table (bool): If True, returns a Polars DataFrame instead of a list of dicts.
        ignore_spectral_data (bool): If True, skips parsing of peak m/z and intensity data.

    Returns:
        list or pl.DataFrame: Parsed MGF blocks.
    """
    # --- Pre-compile regexes (avoid recompilation per line) ---
    _RE_CE_WITH_METHOD = re.compile(r"^(\d+(?:\.\d+)?)\s*(HCD|CID)$", re.IGNORECASE)

    # --- Pre-build lookup sets for O(1) key matching (avoid O(n) list scans) ---
    _COLLISION_ENERGY_KEYS = frozenset({"collision_energy", "ms_ionization_energy"})
    _FRAGMENTATION_KEYS = frozenset({"fragmentation_mode", "ms_frag_mode", "fragmentation_method", "ms_dissociation_method"})
    _IONMODE_KEYS = frozenset({"ionmode", "ion_mode", "ms_ion_mode", "polarity"})
    _FEATURE_ID_KEYS = frozenset({"feature_id", "accession"})
    _ADDUCT_KEYS = frozenset({"precursor_type", "adduct"})
    _PEPMASS_KEYS = frozenset({"precursormz", "precursor_mz", "pepmass", "precursor_mz_value"})
    _INSTRUMENT_KEYS = frozenset(
        {
            "instrument",
            "instrument_model",
            "instrument_model_name",
            "instrument_name",
            "source_instrument",
            "ms_mass_analyzer",
            "instrument_type",
        }
    )
    _NAME_KEYS = frozenset({"name", "compound_name"})
    _MSLEVEL_KEYS = frozenset({"mslevel", "ms_level"})
    _POSITIVE_VALUES = frozenset({"positive", "pos", "p", "+"})
    _NEGATIVE_VALUES = frozenset({"negative", "neg", "n", "-"})

    # --- Build required keys as a frozenset for fast difference operations ---
    required_keys_set = frozenset(
        {
            "pepmass",
            "instrument",
            "name",
            "adduct",
            "ionmode",
            "fragmentation_method",
            "collision_energy",
        }
    )
    if not ignore_spectral_data:
        required_keys_set = required_keys_set | frozenset({"$$SpectrumData"})

    # --- Use plain dicts instead of OrderedDict (Python 3.7+ preserves insertion order) ---
    blocks = []
    blocks_loaded_successfully_n = 0
    current_block_primary = {}
    current_block_secondary = {}
    blocks_not_used = 0

    incomplete_blocks = defaultdict(int)
    cur_id = 0
    filename = os.path.basename(file_path)

    # --- Collect peak lines as raw strings, batch-parse at END IONS ---
    peak_lines = []

    # --- Stream line-by-line instead of readlines() to avoid doubling memory ---
    with open(file_path, "r", errors="ignore") as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue

            if line == "BEGIN IONS":
                current_block_primary = {}
                current_block_secondary = {}
                current_block_secondary["AnnoMe_internal_ID"] = f"{filename}___{cur_id}"
                cur_id += 1
                peak_lines = []

            elif line.lower().startswith("annome_internal_id"):
                pass

            elif line == "END IONS":
                # Batch-parse collected peak lines into spectrum data
                if not ignore_spectral_data and peak_lines:
                    mzs = []
                    intensities = []
                    for pl_line in peak_lines:
                        parts = pl_line.split()
                        mzs.append(float(parts[0]))
                        intensities.append(float(parts[1]))
                    current_block_secondary["$$SpectrumData"] = [mzs, intensities]

                # Track missing keys using frozenset difference (fast, no sorting per block)
                all_block_keys = frozenset(current_block_primary.keys()) | frozenset(current_block_secondary.keys())
                missing = required_keys_set - all_block_keys
                if missing:
                    # Use a sorted tuple as a hashable, stable key for the counter
                    incomplete_blocks[tuple(sorted(missing))] += 1

                # Check if the block has the required keys to be considered valid
                # Inline float check: try/except is only hit on bad data
                pepmass_val = current_block_primary.get("pepmass")
                use_block = False
                if pepmass_val is not None and "name" in current_block_primary:
                    try:
                        float(pepmass_val)
                        use_block = True
                    except (ValueError, TypeError):
                        pass

                if use_block or not check_required_keys:
                    blocks_loaded_successfully_n += 1
                    # Always collect as a merged dict; polars conversion happens at the end
                    blocks.append({**current_block_primary, **current_block_secondary})
                else:
                    blocks_not_used += 1

                # Reset for next block
                current_block_primary = {}
                current_block_secondary = {}
                peak_lines = []

            elif "=" in line:
                key, value = line.split("=", 1)
                key = key.strip().lower()
                value = value.strip()
                is_primary = False

                if key in _COLLISION_ENERGY_KEYS:
                    # Use pre-compiled regex (single match instead of match+match)
                    ce_match = _RE_CE_WITH_METHOD.match(value)
                    if ce_match:
                        value = ce_match.group(1)
                        current_block_primary["fragmentation_method"] = ce_match.group(2).upper()

                    try:
                        if value.startswith("[") and value.endswith("]"):
                            items = sorted(float(v.strip()) for v in value[1:-1].split(",") if v.strip())
                            value = items
                        elif "," in value:
                            items = sorted(float(v.strip()) for v in value.split(",") if v.strip())
                            value = items
                        else:
                            try:
                                value = [round(float(value), 0)]
                            except ValueError:
                                pass
                    except ValueError:
                        pass

                    key = "collision_energy"
                    is_primary = True

                elif key in _FRAGMENTATION_KEYS:
                    key = "fragmentation_method"
                    is_primary = True

                elif key in _IONMODE_KEYS:
                    if key == "polarity":
                        if value not in ("1", "0"):
                            raise ValueError(f"Invalid polarity value: {value}. Use '1' for positive and '0' for negative.")
                        value = "+" if value == "1" else "-"
                    val_lower = value.lower()
                    if val_lower in _POSITIVE_VALUES:
                        value = "+"
                    elif val_lower in _NEGATIVE_VALUES:
                        value = "-"
                    else:
                        value = "NA"
                    key = "ionmode"
                    is_primary = True

                elif key in _FEATURE_ID_KEYS:
                    key = "feature_id"

                elif key in _ADDUCT_KEYS:
                    key = "adduct"
                    is_primary = True

                elif key in _PEPMASS_KEYS:
                    key = "pepmass"
                    is_primary = True

                elif key in _INSTRUMENT_KEYS:
                    key = "instrument"
                    if key in current_block_primary:
                        value = current_block_primary[key] + f"; {value}"
                    is_primary = True

                elif key in _NAME_KEYS:
                    value = value.lower().strip()
                    key = "name"
                    is_primary = True

                elif key in _MSLEVEL_KEYS:
                    key = "MSLEVEL"

                if is_primary:
                    current_block_primary[key] = str(value)
                else:
                    current_block_secondary[key] = str(value)

            elif line.lower().startswith("num peaks"):
                # Just a header line; peak_lines list is already initialized at BEGIN IONS
                pass

            else:
                # Collect raw peak line for batch parsing at END IONS
                if not ignore_spectral_data:
                    peak_lines.append(line)

    # --- Print warnings ---
    if incomplete_blocks:
        print(f"{Fore.RED}")
        print("   - Warning: Some blocks are missing required keys:")
        for keys, count in natsort.natsorted(incomplete_blocks.items(), key=lambda x: str(x[0]).lower()):
            print(f"      - {list(keys)}: {count} blocks")
        print(f"{Style.RESET_ALL}")

    if blocks_not_used > 0:
        print(f"{Fore.RED}")
        print(f"   - Warning: {blocks_not_used} blocks were not used due to missing required keys.")
        print(f"{Style.RESET_ALL}")

    # --- Convert to polars or filter list ---
    if return_as_polars_table:
        # Build polars DataFrame from list of dicts in one shot (avoids O(cols*rows) padding)
        result = pl.DataFrame(blocks)
        assert blocks_loaded_successfully_n == result.shape[0], f"Expected {blocks_loaded_successfully_n} blocks, but got {result.shape[0]}"
        if "MSLEVEL" in result.columns:
            result = result.filter((pl.col("MSLEVEL") != "1") | pl.col("MSLEVEL").is_null())
        return result
    else:
        assert blocks_loaded_successfully_n == len(blocks), f"Expected {blocks_loaded_successfully_n} blocks, but got {len(blocks)}"
        return [block for block in blocks if block.get("MSLEVEL", "2") != "1"]


def standardize_blocks(blocks, standards):
    """Standardizes the blocks by applying standardization functions to specific fields.
    Args:
        blocks (list): List of parsed MGF blocks.
        standards (dict): Dictionary where keys are field names and values are a standardization function or a list of standardization functions.
    Returns:
        None: The function modifies the blocks in place.
    """

    for block in blocks:
        for key in standards.keys():
            if key in block.keys():
                value = block[key]

                if isinstance(standards[key], list):
                    for standard in standards[key]:
                        value = standard(value)
                else:
                    value = standards[key](value)

                # Update the block with the standardized value
                block[key] = value


def export_mgf_file(blocks, output_file_path):
    """
    Exports parsed MGF blocks to a new MGF file.

    Args:
        blocks (dict): Parsed MGF blocks.
        output_file_path (str): Path to the output MGF file.
    """
    with open(output_file_path, "w") as file:
        for feature_blocks in blocks:
            file.write("BEGIN IONS\n")
            for key, value in feature_blocks.items():
                if key == "$$SpectrumData":
                    pass
                elif key.lower() == "collision_energy":
                    file.write(f"{key}={str(value).replace(' ', '')}\n")
                else:
                    file.write(f"{key}={value}\n")
            if "$$SpectrumData" in feature_blocks:
                # file.write("Num peaks {}\n".format(len(feature_blocks["$$SpectrumData"][0])))
                for mzi in range(len(feature_blocks["$$SpectrumData"][0])):
                    file.write(f"{feature_blocks['$$SpectrumData'][0][mzi]} {feature_blocks['$$SpectrumData'][1][mzi]}\n")
            file.write("END IONS\n\n")


def export_mgf_file_from_polars_table(df, output_file_path):
    """
    Exports parsed MGF blocks from a Polars DataFrame to a new MGF file.

    Args:
        df (pl.DataFrame): Parsed MGF blocks in a Polars DataFrame.
        output_file_path (str): Path to the output MGF file.
    """
    with open(output_file_path, "w") as file:
        for row in df.iter_rows(named=True):
            file.write("BEGIN IONS\n")
            for key, value in row.items():
                if key == "$$SpectrumData":
                    pass
                elif key.lower() == "collision_energy":
                    file.write(f"{key}={str(value).replace(' ', '')}\n")
                elif value is not None:
                    file.write(f"{key}={value}\n")
            if "$$SpectrumData" in row and row["$$SpectrumData"] is not None:
                # file.write("Num peaks {}\n".format(len(row["$$SpectrumData"][0])))
                for mzi in range(len(row["$$SpectrumData"][0])):
                    file.write(f"{row['$$SpectrumData'][0][mzi]} {row['$$SpectrumData'][1][mzi]}\n")
            file.write("END IONS\n\n")


def show_overview_of_blocks(blocks):
    """
    Prints an overview of the parsed MGF blocks, including counts of features, instruments, and fragmentation methods.
    Args:
        blocks (list): List of parsed MGF blocks.
    """

    fields = set()
    for block in blocks:
        fields.update(block.keys())

    print(f"   - {len(fields)} unique keys found in blocks: ", end="")
    for i, field in enumerate(natsort.natsorted(fields, key=lambda x: x.lower())):
        print(
            f"{', ' if i > 0 else ''}{Fore.YELLOW if i % 2 == 0 else Fore.GREEN}{field}{Style.RESET_ALL}",
            end="",
        )
    print("")

    instrument_counts = {}
    for block in blocks:
        if "instrument" in block:
            instrument = block["instrument"]
        else:
            instrument = "Unknown"
        instrument_counts[instrument] = instrument_counts.get(instrument, 0) + 1

    if instrument_counts:
        print("   - Instrument usage counts: ", end="")
        sorted_instruments = sorted(instrument_counts.items(), key=lambda item: item[1], reverse=True)
        for i, (instrument, count) in enumerate(sorted_instruments):
            print(
                f"{', ' if i > 0 else ''}{Fore.YELLOW if i % 2 == 0 else Fore.GREEN}{instrument}: {count}{Style.RESET_ALL}",
                end="",
            )
        print("")
    else:
        print("   - No instrument information found in blocks.")

    ionmodes = {}
    for block in blocks:
        if "ionmode" in block:
            ionmode = block["ionmode"]
        else:
            ionmode = "Unknown"
        ionmodes[ionmode] = ionmodes.get(ionmode, 0) + 1

    if ionmodes:
        print("   - Ion mode usage counts: ", end="")
        sorted_ionmodes = sorted(ionmodes.items(), key=lambda item: item[1], reverse=True)
        for i, (ionmode, count) in enumerate(sorted_ionmodes):
            print(
                f"{', ' if i > 0 else ''}{Fore.YELLOW if i % 2 == 0 else Fore.GREEN}{ionmode}: {count}{Style.RESET_ALL}",
                end="",
            )
        print("")
    else:
        print("   - No ion mode information found in blocks.")

    fragmentation_method_counts = {}
    for block in blocks:
        if "fragmentation_method" in block:
            method = block["fragmentation_method"]
        else:
            method = "Unknown"
        fragmentation_method_counts[method] = fragmentation_method_counts.get(method, 0) + 1

    if fragmentation_method_counts:
        print("   - Fragmentation method usage counts: ", end="")
        sorted_methods = sorted(fragmentation_method_counts.items(), key=lambda item: item[1], reverse=True)
        for i, (method, count) in enumerate(sorted_methods):
            print(
                f"{', ' if i > 0 else ''}{Fore.YELLOW if i % 2 == 0 else Fore.GREEN}{method}: {count}{Style.RESET_ALL}",
                end="",
            )
        print()
    else:
        print("   - No fragmentation method information found in blocks.")

    collision_energy_counts = {}
    for block in blocks:
        if "collision_energy" in block:
            energy = block["collision_energy"]
        else:
            energy = "Unknown"
        collision_energy_counts[energy] = collision_energy_counts.get(energy, 0) + 1

    if collision_energy_counts:
        print("   - Collision energy usage counts: ", end="")
        sorted_energies = sorted(collision_energy_counts.items(), key=lambda item: item[1], reverse=True)
        for i, (energy, count) in enumerate(sorted_energies):
            print(
                f"{', ' if i > 0 else ''}{Fore.YELLOW if i % 2 == 0 else Fore.GREEN}{energy}: {count}{Style.RESET_ALL}",
                end="",
            )
        print()
    else:
        print("   - No collision energy information found in blocks.")


def verify_mgf(blocks, smiles_field, name_field):
    """
    Verifies the integrity of MGF blocks by checking for required fields.

    Args:
        blocks (list): List of parsed MGF blocks.

    Returns:
        bool: True if all blocks are valid, False otherwise.
    """

    found = {}

    for block in blocks:
        smiles = None
        name = None

        if smiles_field in block:
            smiles = block[smiles_field]
            if smiles not in found:
                found[smiles] = set()

        if name_field in block:
            name = block[name_field]
            if name not in found:
                found[name] = set()

        if smiles is not None and name is not None:
            found[smiles].add(name.lower().strip())
            found[name].add(smiles.lower().strip())

    error = False
    for smiles, names in found.items():
        if len(names) > 1:
            print(f"Warning: name/smiles '{smiles}' is associated with multiple smiles/names: {', '.join(names)}")
            error = True

    return not error


def get_fields(blocks):
    """
    Extracts the unique fields from the parsed MGF blocks.    Returns:
        set: A set of unique field names found in the blocks.
    """
    fields = set()
    for block in blocks:
        fields.update(block.keys())
    return fields


def filter_blocks_with_required_fields(blocks):
    """
    Filters blocks to check if they contain the necessary fields 'pepmass' and 'rtinseconds'.

    Args:
        blocks (list): List of parsed MGF blocks.

    Returns:
        list: A list of blocks that contain both 'pepmass' and 'rtinseconds' fields.
    """
    filtered_blocks = []
    for block in blocks:
        keys = {key.lower() for key in block.keys()}
        if "pepmass" in keys and "rtinseconds" in keys:
            filtered_blocks.append(block)
    return filtered_blocks


def filter_low_intensity_peaks(blocks, intensity_threshold=0.01):
    """
    Filters out m/z and intensity pairs in the $$SpectrumData field where the intensity
    is less than a specified percentage of the maximum intensity in the block.

    Args:
        blocks (list): List of parsed MGF blocks.
        intensity_threshold (float): The relative intensity threshold (default: 0.01).

    Returns:
        list: A list of blocks with filtered $$SpectrumData.
    """
    for block in blocks:
        if "$$SpectrumData" in block:
            mz, intensity = block["$$SpectrumData"]
            max_intensity = np.max(intensity) if len(intensity) > 0 and np.sum(intensity) > 0 else 0
            if max_intensity > 0:
                use_inds = np.argwhere(intensity >= max_intensity * intensity_threshold).flatten()
                block["$$SpectrumData"] = [mz[use_inds], intensity[use_inds]]
    return blocks


def filter_smiles(smarts_strings, check_fun):
    """
    Filters a list of SMILES strings based on a provided function.
    Args:
        smarts_strings (list): List of SMILES strings to filter.
        check_fun (function): Function that takes a SMILES string and returns True if it matches the criteria, False otherwise.
    Returns:
        tuple: A tuple containing three lists:
            - filtered_smiles: List of SMILES strings that matched the criteria.
            - non_matching_smiles: List of SMILES strings that did not match the criteria.
            - errored_smiles: List of SMILES strings that caused an error during processing.
    """
    filtered_smiles = []
    non_matching_smiles = []
    errored_smiles = []

    for smiles in tqdm(smarts_strings):
        try:
            if check_fun(smiles):
                filtered_smiles.append(smiles)
            else:
                non_matching_smiles.append(smiles)
        except Exception as e:
            print(f"Error processing SMILES '{smiles}': {e}")
            errored_smiles.append(smiles)

    return filtered_smiles, non_matching_smiles, errored_smiles


def draw_smiles(smiles_strings, legends=None, max_draw=10, molsPerRow=10, use_smarts=False):
    """
    Draws a grid of SMILES strings as images.
    Args:
        smiles_strings (list): List of SMILES strings or RDKit Mol objects to draw.
        legends (list, optional): List of legends for each SMILES string.
        max_draw (int): Maximum number of SMILES strings to draw.
        molsPerRow (int): Number of molecules per row in the grid.
    Returns:
        Image: An RDKit image object containing the drawn SMILES strings.
    """

    if type(smiles_strings) is str or type(smiles_strings) is rdkit.Chem.rdchem.Mol:
        smiles_strings = [smiles_strings]

    mols = []
    for s in smiles_strings:
        add = []
        if type(s) is list:
            add.extend(s)
        elif type(s) is rdkit.Chem.rdchem.Mol:
            add.append(s)
        else:
            add = [s]

        for x in add:
            if type(x) == str:
                if use_smarts:
                    mol = rdkit.Chem.MolFromSmarts(x)
                else:
                    mol = rdkit.Chem.MolFromSmiles(x)
            elif type(x) is rdkit.Chem.rdchem.Mol:
                mol = x
            else:
                raise ValueError(f"Unsupported type for SMILES: {type(x)}")
            mols.append(mol)

    if len(smiles_strings) > max_draw:
        selected_indices = np.random.choice(len(smiles_strings), size=max_draw, replace=False)
        selected_mols = [mols[i] for i in selected_indices]
        selected_legends = [legends[i] for i in selected_indices] if legends else None
    else:
        selected_mols = mols
        selected_legends = legends if legends else None

    return rdkit.Chem.Draw.MolsToGridImage(
        selected_mols,
        legends=selected_legends,
        subImgSize=(500, 500),
        molsPerRow=min(molsPerRow, len(selected_mols)),
        maxMols=min(max_draw, len(selected_mols)),
        useSVG=False,
    )


def draw_smarts(smarts_strings, legends=None, max_draw=10, molsPerRow=10):
    """
    Draws a grid of SMARTS strings as images.
    Args:
        smarts_strings (list): List of SMARTS strings to draw.
        legends (list, optional): List of legends for each SMARTS string.
        max_draw (int): Maximum number of SMARTS strings to draw.
        molsPerRow (int): Number of molecules per row in the grid.
    Returns:
        Image: An RDKit image object containing the drawn SMARTS strings.
    """
    return draw_smiles(smarts_strings, legends, max_draw, molsPerRow, use_smarts=True)


def draw_names(compounds, blocks, name_field, smiles_field, max_draw=10, molsPerRow=10):
    """
    Plots the names of compounds as images.

    Args:
        cmpds (set): Set of compound names.
        blocks (list): List of parsed MGF blocks.
        max_draw (int): Maximum number of compounds to draw.
        molsPerRow (int): Number of molecules per row in the grid.
    """
    smiles = []
    legends = []
    for name in compounds:
        for block in blocks:
            if name_field in block.keys() and block[name_field] == name and smiles_field in block.keys() and block[smiles_field]:
                smiles.append(block[smiles_field])
                legends.append(name)
                break

    sorted_indices = sorted(range(len(legends)), key=lambda i: legends[i].lower())
    smiles = [smiles[i] for i in sorted_indices]
    legends = [legends[i] for i in sorted_indices]

    return draw_smiles(smiles, legends, max_draw, molsPerRow)


def substructure_fn(smiles, substructures_to_match=None):
    """
    Checks if a given SMILES string contains all specified substructures.
    Args:
        smiles (str): The SMILES string to check.
        substructures_to_match (list or str or rdkit.Chem.rdchem.Mol, optional):
            A list of substructures to match, or a single substructure as a string or Mol object.
    Returns:
        bool: True if all substructures are found in the SMILES, False otherwise.
    """
    mol = rdkit.Chem.MolFromSmiles(smiles)

    if substructures_to_match is None:
        raise ValueError("No substructures to match provided")

    if type(substructures_to_match) is str or type(substructures_to_match) is rdkit.Chem.rdchem.Mol:
        substructures_to_match = [substructures_to_match]

    for sub in substructures_to_match:
        tomatch = None
        if type(sub) is list:
            tomatch = sub
        else:
            tomatch = [sub]

        anyMatch = False
        for smarts in tomatch:
            if type(smarts) is str:
                sub = rdkit.Chem.MolFromSmarts(smarts)
            elif type(smarts) is rdkit.Chem.rdchem.Mol:
                sub = smarts
            else:
                raise ValueError(f"Unsupported type for substructure: {type(smarts)}")
            anyMatch = anyMatch or mol.HasSubstructMatch(sub)

        if not anyMatch:
            return False
    return True


def prep_smarts_key(smart, replace=True, convert_to_rdkit=True):
    """
    Prepares a SMARTS string for use in RDKit by replacing 'c' with 'C' and 'o' with 'O' (i.e, no carbon or oxygen atom are aromatic or aliphatic).
    Args:
        smart (str): The SMARTS string to prepare.
        replace (bool): If True, replaces 'c' with 'C' and 'o' with 'O' in the SMARTS string.
    Returns:
        rdkit.Chem.rdchem.Mol: The RDKit Mol object corresponding to the prepared SMARTS string.
    """

    replacements = [
        ("[c,C]", "[C,c]"),
        ("[o,O]", "[O,o]"),
        ("[C,c]", "C"),
        ("[O,o]", "O"),
        ("c", "C"),
        ("o", "O"),
        ("C", "[C,c]"),
        ("O", "[O,o]"),
    ]

    if replace:
        for old, new in replacements:
            smart = smart.replace(old, new)
    if not convert_to_rdkit:
        return smart
    else:
        return rdkit.Chem.MolFromSmarts(smart)


def generate_and_save_image(chunk_idx, matching_compounds_list, spectra, name_field, smiles_field, chunk_size, output_folder, database_name, typ):
    chunk_compounds = matching_compounds_list[chunk_idx : chunk_idx + chunk_size]
    try:
        img = draw_names(
            chunk_compounds,
            spectra,
            name_field,
            smiles_field,
            max_draw=chunk_size,
        )
        out_file = f"{output_folder}/{database_name}___{typ}_chunk{chunk_idx}.png"
        with open(out_file, "wb") as f:
            f.write(img.data)
        print(f"   - Exported images {chunk_idx}-{chunk_idx + chunk_size} for substructures to {Fore.YELLOW}{out_file}{Style.RESET_ALL}")
        return out_file
    except Exception as e:
        print(f"ERROR: image generation failed, continuing without plotting substructures: {e}")
    return None


def process_database(
    database_name,
    mgf_file,
    smiles_field,
    name_field,
    sf_field,
    smart_checks,
    standardize_block_functions,
    output_folder,
    include_compound_plots=None,
    filter_fn=None,
    verbose=False,
    parallel_threads_plotting=24,
):
    """
    Processes an MGF file, standardizes the blocks, checks if these contain the necessary substructures, and generates a summary table with compound information.
    Args:
        database_name (str): Name of the database.
        mgf_file (str): Path to the MGF file to process.
        smiles_field (str): Field name for SMILES strings in the MGF blocks.
        name_field (str): Field name for compound names in the MGF blocks.
        sf_field (str): Field name for sum formulas in the MGF blocks.
        smart_checks (list): List of SMARTS strings to check against the SMILES strings.
        standardize_block_functions (dict): Dictionary of functions to standardize specific fields in the MGF blocks.
        output_folder (str): Path to the folder where the output files will be saved.
        include_compound_plots (bool, optional): If True, generates compound structure plots. Defaults to None.
        filter_fn (function, optional): Custom filter function to apply to the blocks. Defaults to
            None.
        verbose (bool, optional): If True, prints additional information during processing. Defaults to False
    Returns:
        tuple: A tuple containing:
            - found_results (dict): Dictionary of found results with compound names as keys.
            - spectra (list): List of processed MGF blocks.
            - generated_files (list): List of generated files during processing.
    """
    start = time.time()

    # keep track of generated results and files
    found_results = {}
    generated_files = []

    # parse MGF file
    spectra = parse_mgf_file(mgf_file)
    print(f"   1. Parsed {Fore.YELLOW}{len(spectra)}{Style.RESET_ALL} blocks")

    # show overview of the spectra
    show_overview_of_blocks(spectra)

    # filter for spectra that have smiles code
    spectra = [block for block in spectra if smiles_field in block.keys() and block[smiles_field] is not None and block[smiles_field].strip().lower() not in ["", "n/a", "na", "none", "null"]]
    # apply custom filter if provided
    if filter_fn is not None:
        spectra = filter_fn(spectra)
    print(f"\n   2. Filtered to {Fore.YELLOW}{len(spectra)}{Style.RESET_ALL} blocks with SMILES string, warning, must not be valid smiles")

    if len(spectra) == 0:
        print(f"\n{Fore.RED}ERROR: No valid spectra found after filtering. {Style.RESET_ALL}")
        return found_results, spectra, generated_files

    # show overview of the spectra
    show_overview_of_blocks(spectra)

    standardize_blocks(spectra, standardize_block_functions)
    print(f"\n   3. standardized block information")

    # show overview of the spectra
    show_overview_of_blocks(spectra)

    # generate final data table for review
    table_data = OrderedDict()

    # get unique names and smiles from the spectra
    names = set([block[name_field] for block in spectra if name_field in block.keys() and block[name_field]])
    print(f"\n   4. Found {Fore.YELLOW}{len(names)}{Style.RESET_ALL} unique compound names in the MGF file")

    temp_dir = tempfile.TemporaryDirectory()

    # add each name to the table data
    if len(names) > 2000:
        if include_compound_plots is None:
            print(
                f"{Fore.YELLOW}WARNING: More than 2000 unique names found, plotting disabled as it might take longer than {len(names) / 50} seconds.\nConsider setting the parameter include_compound_plots to True.{Style.RESET_ALL}"
            )
            include_compound_plots = False
        elif include_compound_plots:
            print(f"{Fore.RED}WARNING: More than 2000 unique names found, plotting may take a long time.{Style.RESET_ALL}")

    smiles = defaultdict(list)
    formulas = defaultdict(list)
    cas = defaultdict(list)
    for block in spectra:
        if name_field in block.keys():
            if smiles_field in block.keys() and block[smiles_field]:
                smiles[block[name_field]].append(block[smiles_field])
            if sf_field in block.keys() and block[sf_field]:
                formulas[block[name_field]].append(block[sf_field])
            if "cas" in block.keys() and block["cas"]:
                cas[block[name_field]].append(block["cas"])
    for i, name in tqdm(
        enumerate(list(natsort.natsorted(names, key=lambda x: x.lower()))),
        position=0,
        leave=False,
        total=len(names),
    ):
        # add compound
        table_data[name] = {"A_name": name, "A_comment": ""}

        # get unique smiles for the compound
        csmiles = set(smiles[name])  # use set to ensure uniqueness
        table_data[name]["A_uniqueSmiles"] = str(csmiles.pop()) if len(csmiles) == 1 else f"{len(csmiles)}: {csmiles}"

        # get unique sum formulas
        cformulas = set(formulas[name])  # use set to ensure uniqueness
        table_data[name]["A_SumFormula"] = str(cformulas.pop()) if len(cformulas) == 1 else f"{len(cformulas)}: {str(cformulas)}"

        # get unique cas numbers
        ccas = set(cas[name])  # use set to ensure uniqueness
        table_data[name]["A_CAS"] = str(ccas.pop()) if len(ccas) == 1 else f"{len(ccas)}: {str(ccas)}"

        # draw structure if possible
        if include_compound_plots is None or include_compound_plots:
            try:
                img = draw_smiles(set(smiles[name]), max_draw=500)
                open(
                    f"{temp_dir.name}/img_{i}.png",
                    "wb",
                ).write(img.data)
                table_data[name]["A_structure"] = f"$$$IMG:{temp_dir.name}/img_{i}.png"
            except Exception as e:
                print(f"ERROR: could not draw structure for {name}: {e}")
                table_data[name]["A_structure"] = "ERROR: could not draw structure"

    n_spectra_with_smiles = sum(1 for block in spectra if smiles_field in block.keys() and block[smiles_field])
    print(f"\n   5. Found {Fore.YELLOW}{n_spectra_with_smiles}{Style.RESET_ALL} spectra with valid SMILES")

    unique_smiles_strings = sorted(list(set([block[smiles_field] for block in spectra if smiles_field in block.keys() and smiles_field != ""])))
    print(f"\n   6. Found {Fore.YELLOW}{len(unique_smiles_strings)}{Style.RESET_ALL} unique smiles strings")

    # process each check, and export matching compounds and spectra
    all_non_matching_smiles = set(unique_smiles_strings)
    chunk_size = 500
    for check_name, subs in smart_checks.items():
        found_results[check_name] = {}

        print("\n--------------------------------------------------------------------------")
        print(f"   # Checking for substructure '{check_name}'")

        matching_smiles, non_matching_smiles, errored_smiles = filter_smiles(unique_smiles_strings, lambda x: substructure_fn(x, subs))
        found_results[check_name]["matching_smiles"] = matching_smiles

        if len(matching_smiles) > 0:
            matching_compounds = set()
            matching_blocks = []
            for spectrum in spectra:
                if smiles_field in spectrum.keys() and spectrum[smiles_field] in matching_smiles:
                    matching_compounds.add(spectrum[name_field])
                    matching_blocks.append(spectrum)

            print(f"   - Found {Fore.YELLOW}{len(matching_compounds)}{Style.RESET_ALL} compounds with {Fore.YELLOW}matching{Style.RESET_ALL} SMILES for {Fore.YELLOW}{check_name}{Style.RESET_ALL}")
            for name in natsort.natsorted(matching_compounds, key=lambda x: x.lower()):
                table_data[name][f"C_{check_name}"] = "substructure match"
                if verbose:
                    print(f"      * {name}")
            found_results[check_name]["matching_compounds"] = matching_compounds

            out_file = f"{output_folder}/{database_name}___{check_name}__MatchingSmiles.mgf"
            export_mgf_file(matching_blocks, out_file)
            print(f"   - Exported spectra to {Fore.YELLOW}{out_file}{Style.RESET_ALL}")
            generated_files.append(out_file)

            # Partition matching_compounds and spectra into chunks of 500
            matching_compounds_list = list(natsort.natsorted(matching_compounds, key=lambda x: x.lower()))

            if include_compound_plots is None or include_compound_plots:
                with ThreadPoolExecutor(max_workers=parallel_threads_plotting) as executor:
                    futures = {
                        executor.submit(
                            generate_and_save_image, chunk_idx, matching_compounds_list, spectra, name_field, smiles_field, chunk_size, output_folder, database_name, f"{check_name}__MatchingSmiles"
                        ): chunk_idx
                        for chunk_idx in range(0, len(matching_compounds_list), chunk_size)
                    }
                    for future in as_completed(futures):
                        out_file = future.result()
                        if out_file:
                            generated_files.append(out_file)

        else:
            print("   - No matches found")

        all_non_matching_smiles = all_non_matching_smiles.intersection(set(non_matching_smiles))

    # export all compounds and structures that did not match any of the substructure checks
    if len(all_non_matching_smiles) > 0:
        matching_compounds = set()
        matching_blocks = []
        for spectrum in spectra:
            if smiles_field in spectrum.keys() and spectrum[smiles_field] in all_non_matching_smiles:
                matching_compounds.add(spectrum[name_field])
                matching_blocks.append(spectrum)

        print(f"   - Found {Fore.YELLOW}{len(matching_compounds)}{Style.RESET_ALL} compounds not {Fore.YELLOW}matching{Style.RESET_ALL} and substructure filter")

        out_file = f"{output_folder}/{database_name}___NonMatchingSmiles.mgf"
        export_mgf_file(matching_blocks, out_file)
        print(f"   - Exported spectra to {Fore.YELLOW}{out_file}{Style.RESET_ALL}")
        generated_files.append(out_file)

        # Partition matching_compounds and spectra into chunks of 500
        matching_compounds_list = list(natsort.natsorted(matching_compounds, key=lambda x: x.lower()))
        # Parallelize image generation for chunks of non-matching compounds

        if include_compound_plots is None or include_compound_plots:
            with ThreadPoolExecutor(max_workers=parallel_threads_plotting) as executor:
                futures = {
                    executor.submit(
                        generate_and_save_image, chunk_idx, matching_compounds_list, spectra, name_field, smiles_field, chunk_size, output_folder, database_name, "NonMatchingSmiles"
                    ): chunk_idx
                    for chunk_idx in range(0, len(matching_compounds_list), chunk_size)
                }
                for future in as_completed(futures):
                    out_file = future.result()
                    if out_file:
                        generated_files.append(out_file)

    # write the table to an Excel file
    out_file = f"{output_folder}/{database_name}___table.xlsx"
    list_to_excel_table([v for k, v in table_data.items()], out_file, column_width=40, row_height=40)
    generated_files.append(out_file)
    temp_dir.cleanup()

    print(f"   - Exported table to {Fore.YELLOW}{out_file}{Style.RESET_ALL}")
    end = time.time()
    print(f"   - Processed {Fore.YELLOW}{len(spectra)}{Style.RESET_ALL} spectra in {end - start:.2f} seconds")
    print("--------------------------------------------------------------------------\n")

    return found_results, spectra, generated_files
